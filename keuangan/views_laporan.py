from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseBadRequest
from .models import Jurnal, Akun
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from .pdf_utils import (
    export_buku_besar_pdf, 
    export_neraca_saldo_pdf, 
    export_laporan_keuangan_pdf,
    create_pdf_response
)


# Shared helper to compute balances and totals by account type for a user
def compute_account_totals_for_user(user):
    akun_list = Akun.objects.filter(user=user)
    saldo = []
    total_debit = 0
    total_kredit = 0

    # Totals by account type
    total_aktiva = 0
    total_kewajiban = 0
    total_modal = 0
    total_pendapatan = 0
    total_beban = 0

    # Define account type groups and their normal balances
    asset_types = ['Aset', 'Aktiva', 'Aktiva Lancar', 'Aktiva Tetap', 'Aktiva Lainnya']
    liability_types = ['Kewajiban', 'Kewajiban Lancar', 'Kewajiban Jangka Panjang']

    for akun in akun_list:
        debit_sum = Jurnal.objects.filter(akun=akun).aggregate(total_debit=Coalesce(Sum('debit'), Value(0, output_field=DecimalField())))['total_debit']
        kredit_sum = Jurnal.objects.filter(akun=akun).aggregate(total_kredit=Coalesce(Sum('kredit'), Value(0, output_field=DecimalField())))['total_kredit']
        saldo_akhir = debit_sum - kredit_sum

        saldo.append({
            'akun': akun,
            'debit': debit_sum,
            'kredit': kredit_sum,
            'saldo': saldo_akhir,
        })

        total_debit += debit_sum
        total_kredit += kredit_sum

        # Determine normal side by account type
        if akun.tipe in asset_types or akun.tipe == 'Beban':
            normal = 'debit'
        else:
            # treat Kewajiban, Modal, Pendapatan as credit-normal
            normal = 'credit'

        # Route balances to correct totals respecting normal balance
        # If account has a debit-normal balance (debit_sum - kredit_sum >= 0)
        if normal == 'debit':
            if saldo_akhir >= 0:
                # debit balance -> goes to left side (aktiva or beban)
                if akun.tipe in asset_types:
                    total_aktiva += saldo_akhir
                elif akun.tipe == 'Beban':
                    total_beban += saldo_akhir
                else:
                    # default to aktiva if unknown
                    total_aktiva += saldo_akhir
            else:
                # credit balance on a debit-normal account -> count on right side
                if akun.tipe in asset_types:
                    total_kewajiban += abs(saldo_akhir)
                elif akun.tipe == 'Beban':
                    total_pendapatan += abs(saldo_akhir)
                else:
                    total_kewajiban += abs(saldo_akhir)
        else:
            # credit-normal accounts
            if saldo_akhir <= 0:
                val = abs(saldo_akhir)
                if akun.tipe in liability_types:
                    total_kewajiban += val
                elif akun.tipe == 'Modal':
                    total_modal += val
                elif akun.tipe == 'Pendapatan':
                    total_pendapatan += val
                else:
                    total_kewajiban += val
            else:
                # unexpected debit balance on credit-normal account -> treat as aktiva
                total_aktiva += saldo_akhir

    return {
        'saldo': saldo,
        'total_debit': total_debit,
        'total_kredit': total_kredit,
        'total_aktiva': total_aktiva,
        'total_kewajiban': total_kewajiban,
        'total_modal': total_modal,
        'total_pendapatan': total_pendapatan,
        'total_beban': total_beban,
    }

@login_required
def buku_besar(request):
    # Ambil parameter filter dari query string
    akun_id = request.GET.get('akun_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Daftar akun milik user untuk dropdown
    akun_list = Akun.objects.filter(user=request.user).order_by('kode')

    # Base queryset: hanya jurnal milik akun user yang login
    base_qs = Jurnal.objects.select_related('transaksi', 'akun').filter(akun__user=request.user)

    # Terapkan filter akun jika dipilih
    if akun_id:
        try:
            akun_filter = Akun.objects.get(id=akun_id, user=request.user)
            base_qs = base_qs.filter(akun=akun_filter)
        except Akun.DoesNotExist:
            akun_filter = None
    else:
        akun_filter = None

    # Terapkan filter tanggal jika ada (robust parsing: YYYY-MM-DD or MM/DD/YYYY)
    from django.utils.dateparse import parse_date as django_parse_date
    from datetime import datetime

    def parse_date_any(s):
        if not s:
            return None
        d = django_parse_date(s)
        if d:
            return d
        try:
            return datetime.strptime(s, '%m/%d/%Y').date()
        except ValueError:
            return None

    sd = parse_date_any(start_date)
    ed = parse_date_any(end_date)
    if sd:
        base_qs = base_qs.filter(transaksi__tanggal__gte=sd)
    if ed:
        base_qs = base_qs.filter(transaksi__tanggal__lte=ed)

    # Urutkan secara default
    if akun_filter:
        qs = base_qs.order_by('transaksi__tanggal', 'transaksi__id')
    else:
        qs = base_qs.order_by('akun__kode', 'transaksi__tanggal', 'transaksi__id')

    # Bentuk data per-akun sesuai filter aktif
    data = []
    grand_total_debit = 0
    grand_total_kredit = 0

    # Jika spesifik akun, hanya satu blok data
    if akun_filter:
        jurnal = list(qs)
        total_debit = sum([j.debit or 0 for j in jurnal])
        total_kredit = sum([j.kredit or 0 for j in jurnal])
        data.append({'akun': akun_filter, 'jurnal': jurnal, 'total_debit': total_debit, 'total_kredit': total_kredit})
        grand_total_debit += total_debit
        grand_total_kredit += total_kredit
    else:
        # Kelompokkan berdasarkan akun
        from collections import defaultdict
        grouped = defaultdict(list)
        for j in qs:
            grouped[j.akun].append(j)
        for akun in akun_list:
            jurnal = grouped.get(akun, [])
            total_debit = sum([j.debit or 0 for j in jurnal])
            total_kredit = sum([j.kredit or 0 for j in jurnal])
            data.append({'akun': akun, 'jurnal': jurnal, 'total_debit': total_debit, 'total_kredit': total_kredit})
            grand_total_debit += total_debit
            grand_total_kredit += total_kredit

    context = {
        'data': data,
        'all_akun': akun_list,
        'grand_total_debit': grand_total_debit,
        'grand_total_kredit': grand_total_kredit,
    }

    return render(request, 'buku_besar.html', context)

@login_required
def neraca_saldo(request):
    # Filters: start_date, end_date, jenis_akun
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    # Accept either 'jenis' (template) or 'jenis_akun' (fallback)
    jenis_akun = request.GET.get('jenis') or request.GET.get('jenis_akun')  # e.g., 'Aktiva', 'Kewajiban', 'Modal', 'Pendapatan', 'Beban'

    from django.utils.dateparse import parse_date as django_parse_date
    from datetime import datetime

    def parse_date_any(s):
        if not s:
            return None
        d = django_parse_date(s)
        if d:
            return d
        try:
            return datetime.strptime(s, '%m/%d/%Y').date()
        except ValueError:
            return None

    sd = parse_date_any(start_date)
    ed = parse_date_any(end_date)

    # Akun list filtered by user and optional jenis
    akun_qs = Akun.objects.filter(user=request.user).order_by('kode')
    if jenis_akun:
        akun_qs = akun_qs.filter(tipe=jenis_akun)

    # Aggregate per akun within the period
    asset_types = ['Aset', 'Aktiva', 'Aktiva Lancar', 'Aktiva Tetap', 'Aktiva Lainnya']
    liability_types = ['Kewajiban', 'Kewajiban Lancar', 'Kewajiban Jangka Panjang']

    saldo = []
    total_debit_all = 0
    total_kredit_all = 0
    total_aktiva = 0
    total_kewajiban = 0
    total_modal = 0
    total_pendapatan = 0
    total_beban = 0

    for akun in akun_qs:
        j_qs = Jurnal.objects.filter(akun=akun)
        if sd:
            j_qs = j_qs.filter(transaksi__tanggal__gte=sd)
        if ed:
            j_qs = j_qs.filter(transaksi__tanggal__lte=ed)

        debit_sum = j_qs.aggregate(total_debit=Coalesce(Sum('debit'), Value(0, output_field=DecimalField())))['total_debit'] or 0
        kredit_sum = j_qs.aggregate(total_kredit=Coalesce(Sum('kredit'), Value(0, output_field=DecimalField())))['total_kredit'] or 0

        total_debit_all += debit_sum
        total_kredit_all += kredit_sum

        # Display columns: split into debet/kredit balance
        if debit_sum >= kredit_sum:
            disp_debet = debit_sum - kredit_sum
            disp_kredit = 0
        else:
            disp_debet = 0
            disp_kredit = kredit_sum - debit_sum

        # Build 'saldo' entries to match existing template expectations
        saldo.append({
            'akun': akun,
            'debit': disp_debet,
            'kredit': disp_kredit,
        })

        # Route saldo akhir to totals by normal balance
        saldo_akhir = debit_sum - kredit_sum
        if akun.tipe in asset_types or akun.tipe == 'Beban':
            # debit-normal
            if saldo_akhir >= 0:
                if akun.tipe in asset_types:
                    total_aktiva += saldo_akhir
                elif akun.tipe == 'Beban':
                    total_beban += saldo_akhir
            else:
                # credit balance on debit-normal
                if akun.tipe in asset_types:
                    total_kewajiban += abs(saldo_akhir)
                elif akun.tipe == 'Beban':
                    total_pendapatan += abs(saldo_akhir)
        else:
            # credit-normal (Kewajiban, Modal, Pendapatan)
            if saldo_akhir <= 0:
                val = abs(saldo_akhir)
                if akun.tipe in liability_types:
                    total_kewajiban += val
                elif akun.tipe == 'Modal':
                    total_modal += val
                elif akun.tipe == 'Pendapatan':
                    total_pendapatan += val
            else:
                # unexpected debit balance -> count to aktiva
                total_aktiva += saldo_akhir

    context = {
        'saldo': saldo,
        'total_debit': total_debit_all,
        'total_kredit': total_kredit_all,
        'total_aktiva': total_aktiva,
        'total_kewajiban': total_kewajiban,
        'total_modal': total_modal,
        'total_pendapatan': total_pendapatan,
        'total_beban': total_beban,
        'laba_rugi': (total_pendapatan - total_beban),
        'rhs_total': (total_kewajiban + total_modal + (total_pendapatan - total_beban)),
    }

    return render(request, 'neraca_saldo.html', context)

@login_required
def laporan_keuangan(request):
    # Delegate to the canonical helper so totals/signs are consistent with neraca
    data = compute_account_totals_for_user(request.user)

    pendapatan = []
    beban = []
    laporan = []
    aktiva = []
    kewajiban = []
    modal = []

    for entry in data['saldo']:
        akun = entry['akun']
        saldo_akhir = entry['saldo']
        laporan.append({'akun': akun, 'saldo': saldo_akhir})

        if akun.tipe in ['Aset', 'Aktiva', 'Aktiva Lancar', 'Aktiva Tetap']:
            aktiva.append({'akun': akun, 'saldo': abs(saldo_akhir) if saldo_akhir > 0 else 0})
        elif akun.tipe in ['Kewajiban', 'Kewajiban Lancar', 'Kewajiban Jangka Panjang']:
            kewajiban.append({'akun': akun, 'saldo': abs(saldo_akhir) if saldo_akhir < 0 else 0})
        elif akun.tipe == 'Modal':
            modal.append({'akun': akun, 'saldo': abs(saldo_akhir) if saldo_akhir < 0 else 0})
        elif akun.tipe == 'Pendapatan':
            pendapatan.append({'akun': akun, 'saldo': abs(saldo_akhir) if saldo_akhir < 0 else 0})
        elif akun.tipe == 'Beban':
            beban.append({'akun': akun, 'saldo': abs(saldo_akhir) if saldo_akhir > 0 else 0})

    total_aktiva = data.get('total_aktiva', 0)
    total_kewajiban = data.get('total_kewajiban', 0)
    total_modal = data.get('total_modal', 0)
    total_pendapatan = data.get('total_pendapatan', 0)
    total_beban = data.get('total_beban', 0)

    # Net income (laba/rugi bersih)
    net_income = total_pendapatan - total_beban

    # Add Retained Earnings / Current Year Earnings into Modal display section
    if net_income != 0:
        modal.append({'akun': Akun(nama='Laba Tahun Berjalan', tipe='Modal'), 'saldo': abs(net_income) if net_income < 0 else net_income})

    # Modal akhir (untuk tampilan): Modal Awal + Laba Bersih
    total_modal_display = total_modal + net_income

    # Accounting equation right side uses Modal Akhir
    equation_right = total_kewajiban + total_modal_display

    # Simple cash flow reconstruction (placeholder logic)
    kas_masuk_operasional = total_pendapatan
    kas_keluar_operasional = total_beban
    kas_bersih_operasional = kas_masuk_operasional - kas_keluar_operasional
    kas_masuk_investasi = 0
    kas_keluar_investasi = 0
    kas_bersih_investasi = kas_masuk_investasi - kas_keluar_investasi
    kas_masuk_pendanaan = total_modal  # treat modal additions as financing inflow
    kas_keluar_pendanaan = 0
    kas_bersih_pendanaan = kas_masuk_pendanaan - kas_keluar_pendanaan
    total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi + kas_bersih_pendanaan
    kas_awal = 0
    kas_akhir = total_aktiva  # Assuming all aktiva currently represented in kas for this simplified model

    # Ratios
    profit_margin = (net_income / total_pendapatan * 100) if total_pendapatan else 0
    current_ratio = (total_aktiva / total_kewajiban) if total_kewajiban else None

    context = {
        'laporan': laporan,
        'aktiva': aktiva,
        'kewajiban': kewajiban,
        'modal': modal,
        'pendapatan': pendapatan,
        'beban': beban,
        'total_aktiva': total_aktiva,
        'total_kewajiban': total_kewajiban,
        'total_modal': total_modal,
        'total_modal_display': total_modal_display,
        'total_pendapatan': total_pendapatan,
        'total_beban': total_beban,
        'net_income': net_income,
        'equation_right': equation_right,
        'kas_masuk_operasional': kas_masuk_operasional,
        'kas_keluar_operasional': kas_keluar_operasional,
        'kas_bersih_operasional': kas_bersih_operasional,
        'kas_masuk_investasi': kas_masuk_investasi,
        'kas_keluar_investasi': kas_keluar_investasi,
        'kas_bersih_investasi': kas_bersih_investasi,
        'kas_masuk_pendanaan': kas_masuk_pendanaan,
        'kas_keluar_pendanaan': kas_keluar_pendanaan,
        'kas_bersih_pendanaan': kas_bersih_pendanaan,
        'total_kas_bersih': total_kas_bersih,
        'kas_awal': kas_awal,
        'kas_akhir': kas_akhir,
        'profit_margin': profit_margin,
        'current_ratio': current_ratio,
    }

    return render(request, 'laporan_keuangan.html', context)

@login_required
def buku_besar_pdf(request):
    """Export Buku Besar to PDF - Only user's data"""
    akun_id = request.GET.get('akun_id')
    akun_name = "Semua Akun"
    
    if akun_id:
        # Filter akun berdasarkan user dan ID
        akun_filter = Akun.objects.get(id=akun_id, user=request.user)
        jurnal = Jurnal.objects.select_related('transaksi', 'akun').filter(akun=akun_filter).order_by('transaksi__tanggal')
        akun_name = f"{akun_filter.kode} - {akun_filter.nama}"
    else:
        # Filter semua jurnal hanya untuk akun milik user yang login
        user_akun_ids = Akun.objects.filter(user=request.user).values_list('id', flat=True)
        jurnal = Jurnal.objects.select_related('transaksi', 'akun').filter(akun__in=user_akun_ids).order_by('akun__kode', 'transaksi__tanggal')
    
    buffer = export_buku_besar_pdf(jurnal, akun_name)
    response = create_pdf_response('buku_besar.pdf')
    response.write(buffer.getvalue())
    
    return response

# --- Export Buku Besar ke Excel ---
from io import BytesIO
@login_required
def buku_besar_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    akun_id = request.GET.get('akun_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    base_qs = Jurnal.objects.select_related('transaksi', 'akun').filter(akun__user=request.user)
    if akun_id:
        try:
            akun_filter = Akun.objects.get(id=akun_id, user=request.user)
            jurnal = base_qs.filter(akun=akun_filter)
            from django.utils.dateparse import parse_date as django_parse_date
            if start_date:
                sd = django_parse_date(start_date)
                if sd:
                    jurnal = jurnal.filter(transaksi__tanggal__gte=sd)
            if end_date:
                ed = django_parse_date(end_date)
                if ed:
                    jurnal = jurnal.filter(transaksi__tanggal__lte=ed)
            jurnal = jurnal.order_by('transaksi__tanggal')
        except Akun.DoesNotExist:
            return HttpResponseBadRequest("Akun tidak valid untuk Excel.")
    else:
        jurnal = base_qs
        from django.utils.dateparse import parse_date as django_parse_date
        if start_date:
            sd = django_parse_date(start_date)
            if sd:
                jurnal = jurnal.filter(transaksi__tanggal__gte=sd)
        if end_date:
            ed = django_parse_date(end_date)
            if ed:
                jurnal = jurnal.filter(transaksi__tanggal__lte=ed)
        jurnal = jurnal.order_by('akun__kode', 'transaksi__tanggal')

    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Besar"

    headers = ["No", "Tanggal", "Kode Akun", "Nama Akun", "Keterangan", "Debet", "Kredit"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563eb")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx != 1 else 6

    total_debet = 0
    total_kredit = 0
    for idx, j in enumerate(jurnal, 1):
        ws.append([
            idx,
            j.transaksi.tanggal.strftime('%d-%m-%Y') if hasattr(j.transaksi, 'tanggal') else '',
            j.akun.kode,
            j.akun.nama,
            getattr(j.transaksi, 'deskripsi', ''),
            float(j.debit) if j.debit else 0,
            float(j.kredit) if j.kredit else 0,
        ])
        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=idx+1, column=col_num)
            cell.border = border
            cell.alignment = right if col_num in [6,7] else center
        total_debet += float(j.debit) if j.debit else 0
        total_kredit += float(j.kredit) if j.kredit else 0

    total_row = len(jurnal) + 2
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_debet).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_kredit).font = Font(bold=True)
    for col_num in range(1, len(headers)+1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.border = border
        cell.alignment = center

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=buku_besar.xlsx'
    return response

@login_required
def neraca_saldo_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Use same filtering logic as neraca_saldo view
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    jenis_akun = request.GET.get('jenis') or request.GET.get('jenis_akun')

    from django.utils.dateparse import parse_date as django_parse_date
    from datetime import datetime

    def parse_date_any(s):
        if not s:
            return None
        d = django_parse_date(s)
        if d:
            return d
        try:
            return datetime.strptime(s, '%m/%d/%Y').date()
        except ValueError:
            return None

    sd = parse_date_any(start_date)
    ed = parse_date_any(end_date)

    akun_qs = Akun.objects.filter(user=request.user).order_by('kode')
    if jenis_akun:
        akun_qs = akun_qs.filter(tipe=jenis_akun)

    wb = Workbook()
    ws = wb.active
    ws.title = "Neraca Saldo"

    # Title & subtitle
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value="NERACA SALDO")
    t.font = Font(size=14, bold=True)
    t.alignment = Alignment(horizontal="center")
    subtitle = []
    if sd or ed:
        sd_disp = start_date or "(awal)"
        ed_disp = end_date or "(akhir)"
        subtitle.append(f"Periode: {sd_disp} s/d {ed_disp}")
    if jenis_akun:
        subtitle.append(f"Jenis: {jenis_akun}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    s = ws.cell(row=2, column=1, value=" | ".join(subtitle))
    s.alignment = Alignment(horizontal="center")

    headers = ["#", "Kode Akun", "Akun", "Debet", "Kredit"]
    ws.append(headers)
    header_row = 3
    header_fill = PatternFill("solid", fgColor="2563eb")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    widths = {1:6, 2:12, 3:32, 4:16, 5:16}
    for col in range(1, 6):
        c = ws.cell(row=header_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = widths[col]

    total_debet_all = 0
    total_kredit_all = 0
    currency = "#,##0"

    row_idx = header_row
    for idx, akun in enumerate(akun_qs, 1):
        j_qs = Jurnal.objects.filter(akun=akun)
        if sd:
            j_qs = j_qs.filter(transaksi__tanggal__gte=sd)
        if ed:
            j_qs = j_qs.filter(transaksi__tanggal__lte=ed)
        debit_sum = j_qs.aggregate(total_debit=Coalesce(Sum('debit'), Value(0, output_field=DecimalField())))['total_debit'] or 0
        kredit_sum = j_qs.aggregate(total_kredit=Coalesce(Sum('kredit'), Value(0, output_field=DecimalField())))['total_kredit'] or 0
        disp_debet = debit_sum - kredit_sum if debit_sum >= kredit_sum else 0
        disp_kredit = kredit_sum - debit_sum if kredit_sum > debit_sum else 0

        row_idx += 1
        ws.append([idx, akun.kode, akun.nama, disp_debet, disp_kredit])
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = right if col in [4,5] else center
            if col in [4,5]:
                cell.number_format = currency

        total_debet_all += disp_debet
        total_kredit_all += disp_kredit

    # Totals
    total_row = row_idx + 1
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    td = ws.cell(row=total_row, column=4, value=total_debet_all)
    tk = ws.cell(row=total_row, column=5, value=total_kredit_all)
    td.font = Font(bold=True)
    tk.font = Font(bold=True)
    td.number_format = currency
    tk.number_format = currency
    for col in range(1, 6):
        c = ws.cell(row=total_row, column=col)
        c.border = border
        c.alignment = right if col in [4,5] else center

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=neraca_saldo.xlsx'
    return resp

@login_required
def laporan_keuangan_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Period filters (for subtitle display)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Reuse compute_account_totals_for_user for totals
    data = compute_account_totals_for_user(request.user)

    wb = Workbook()
    ws_bs = wb.active
    ws_bs.title = "Neraca"

    # Balance Sheet
    ws_bs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_bs = ws_bs.cell(row=1, column=1, value="NERACA")
    title_bs.font = Font(bold=True, size=14)
    title_bs.alignment = Alignment(horizontal="center")
    # Subtitle with period
    ws_bs.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    subtitle_bs = ws_bs.cell(row=2, column=1, value=f"Periode: {start_date or '(awal)'} - {end_date or '(akhir)'}")
    subtitle_bs.alignment = Alignment(horizontal="center")
    headers_bs = ["Kategori", "Akun", "Jumlah (Rp)"]
    ws_bs.append(headers_bs)
    header_fill = PatternFill("solid", fgColor="2563eb")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Header styling (row 3 due to subtitle)
    for col in range(1, 4):
        c = ws_bs.cell(row=3, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_bs.column_dimensions[get_column_letter(col)].width = [14, 32, 18][col-1]

    # Aktiva
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    start_row = ws_bs.max_row
    for entry in data['saldo']:
        akun = entry['akun']
        saldo = entry['saldo']
        if akun.tipe in ['Aset', 'Aktiva', 'Aktiva Lancar', 'Aktiva Tetap', 'Aktiva Lainnya'] and saldo != 0:
            ws_bs.append(["Aktiva", akun.nama, abs(saldo) if saldo > 0 else 0])
            for col in range(1, 4):
                cell = ws_bs.cell(row=ws_bs.max_row, column=col)
                cell.border = border
                cell.alignment = right if col == 3 else Alignment(horizontal="left")
                if col == 3:
                    cell.number_format = "#,##0"
            # Zebra striping
            if (ws_bs.max_row - start_row) % 2 == 0:
                for col in range(1, 4):
                    ws_bs.cell(row=ws_bs.max_row, column=col).fill = zebra_fill
    ws_bs.append(["Total Aktiva", "", data.get('total_aktiva', 0)])
    for col in range(1, 4):
        cell = ws_bs.cell(row=ws_bs.max_row, column=col)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = right if col == 3 else Alignment(horizontal="left")
        if col == 3:
            cell.number_format = "#,##0"

    # Kewajiban & Modal
    start_row_km = ws_bs.max_row
    for entry in data['saldo']:
        akun = entry['akun']
        saldo = entry['saldo']
        if akun.tipe in ['Kewajiban', 'Kewajiban Lancar', 'Kewajiban Jangka Panjang'] and saldo != 0:
            ws_bs.append(["Kewajiban", akun.nama, abs(saldo) if saldo < 0 else 0])
            for col in range(1, 4):
                cell = ws_bs.cell(row=ws_bs.max_row, column=col)
                cell.border = border
                cell.alignment = right if col == 3 else Alignment(horizontal="left")
                if col == 3:
                    cell.number_format = "#,##0"
            if (ws_bs.max_row - start_row_km) % 2 == 0:
                for col in range(1, 4):
                    ws_bs.cell(row=ws_bs.max_row, column=col).fill = zebra_fill
    for entry in data['saldo']:
        akun = entry['akun']
        saldo = entry['saldo']
        if akun.tipe == 'Modal' and saldo != 0:
            ws_bs.append(["Modal", akun.nama, abs(saldo) if saldo < 0 else 0])
            for col in range(1, 4):
                cell = ws_bs.cell(row=ws_bs.max_row, column=col)
                cell.border = border
                cell.alignment = right if col == 3 else Alignment(horizontal="left")
                if col == 3:
                    cell.number_format = "#,##0"
    ws_bs.append(["Total Kewajiban + Modal", "", data.get('total_kewajiban', 0) + data.get('total_modal', 0) + (data.get('total_pendapatan', 0) - data.get('total_beban', 0))])
    for col in range(1, 4):
        cell = ws_bs.cell(row=ws_bs.max_row, column=col)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = right if col == 3 else Alignment(horizontal="left")
        if col == 3:
            cell.number_format = "#,##0"

    # Freeze panes
    ws_bs.freeze_panes = ws_bs["A4"]

    # Laba Rugi sheet
    ws_pl = wb.create_sheet("Laba Rugi")
    # Title & subtitle
    ws_pl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t_pl = ws_pl.cell(row=1, column=1, value="LAPORAN LABA RUGI")
    t_pl.font = Font(bold=True, size=14)
    t_pl.alignment = Alignment(horizontal="center")
    ws_pl.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    s_pl = ws_pl.cell(row=2, column=1, value=f"Periode: {start_date or '(awal)'} - {end_date or '(akhir)'}")
    s_pl.alignment = Alignment(horizontal="center")
    ws_pl.append(["Kategori", "Akun", "Jumlah (Rp)"])
    for col in range(1, 4):
        c = ws_pl.cell(row=3, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_pl.column_dimensions[get_column_letter(col)].width = [14, 32, 18][col-1]

    start_row_pl = ws_pl.max_row
    for entry in data['saldo']:
        akun = entry['akun']
        saldo = entry['saldo']
        if akun.tipe == 'Pendapatan' and saldo != 0:
            ws_pl.append(["Pendapatan", akun.nama, abs(saldo) if saldo < 0 else 0])
        elif akun.tipe == 'Beban' and saldo != 0:
            ws_pl.append(["Beban", akun.nama, abs(saldo) if saldo > 0 else 0])
        for col in range(1, 3+1):
            cell = ws_pl.cell(row=ws_pl.max_row, column=col)
            cell.border = border
            cell.alignment = right if col == 3 else Alignment(horizontal="left")
            if col == 3:
                cell.number_format = "#,##0"
        if (ws_pl.max_row - start_row_pl) % 2 == 0:
            for col in range(1, 4):
                ws_pl.cell(row=ws_pl.max_row, column=col).fill = zebra_fill
    ws_pl.append(["Laba/Rugi Bersih", "", data.get('total_pendapatan', 0) - data.get('total_beban', 0)])
    for col in range(1, 4):
        cell = ws_pl.cell(row=ws_pl.max_row, column=col)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = right if col == 3 else Alignment(horizontal="left")
        if col == 3:
            cell.number_format = "#,##0"
    ws_pl.freeze_panes = ws_pl["A4"]

    # Arus Kas sheet
    ws_cf = wb.create_sheet("Arus Kas")
    ws_cf.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t_cf = ws_cf.cell(row=1, column=1, value="LAPORAN ARUS KAS")
    t_cf.font = Font(bold=True, size=14)
    t_cf.alignment = Alignment(horizontal="center")
    ws_cf.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    s_cf = ws_cf.cell(row=2, column=1, value=f"Periode: {start_date or '(awal)'} - {end_date or '(akhir)'}")
    s_cf.alignment = Alignment(horizontal="center")
    ws_cf.append(["Kategori", "Keterangan", "Jumlah (Rp)"])
    for col in range(1, 4):
        c = ws_cf.cell(row=3, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_cf.column_dimensions[get_column_letter(col)].width = [16, 28, 18][col-1]

    cf_rows = [
        ("Operasional", "Kas Masuk", data.get('total_pendapatan', 0)),
        ("Operasional", "Kas Keluar", data.get('total_beban', 0)),
        ("Operasional", "Kas Bersih", data.get('total_pendapatan', 0) - data.get('total_beban', 0)),
        ("Investasi", "Kas Bersih", 0),
        ("Pendanaan", "Kas Masuk Modal", data.get('total_modal', 0)),
        ("Pendanaan", "Kas Bersih", data.get('total_modal', 0)),
    ]
    for r in cf_rows:
        ws_cf.append(list(r))
        for col in range(1, 4):
            cell = ws_cf.cell(row=ws_cf.max_row, column=col)
            cell.border = border
            cell.alignment = right if col == 3 else Alignment(horizontal="left")
            if col == 3:
                cell.number_format = "#,##0"
        if (ws_cf.max_row - 3) % 2 == 0:
            for col in range(1, 4):
                ws_cf.cell(row=ws_cf.max_row, column=col).fill = zebra_fill
    ws_cf.freeze_panes = ws_cf["A4"]

    # Analisis sheet
    ws_an = wb.create_sheet("Analisis")
    ws_an.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    t_an = ws_an.cell(row=1, column=1, value="ANALISIS")
    t_an.font = Font(bold=True, size=14)
    t_an.alignment = Alignment(horizontal="center")
    ws_an.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    s_an = ws_an.cell(row=2, column=1, value=f"Periode: {start_date or '(awal)'} - {end_date or '(akhir)'}")
    s_an.alignment = Alignment(horizontal="center")
    ws_an.append(["Metrix", "Nilai"])
    for col in range(1, 3):
        c = ws_an.cell(row=3, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws_an.column_dimensions[get_column_letter(col)].width = [24, 18][col-1]

    pm = (data.get('total_pendapatan', 0) - data.get('total_beban', 0))
    pm_ratio = (pm / data.get('total_pendapatan', 1) * 100) if data.get('total_pendapatan', 0) else 0
    cr_ratio = (data.get('total_aktiva', 0) / data.get('total_kewajiban', 1)) if data.get('total_kewajiban', 0) else 0
    ws_an.append(["Profit Margin (%)", pm_ratio])
    ws_an.append(["Current Ratio", cr_ratio])
    ws_an['B4'].number_format = "0.00%"
    ws_an['B5'].number_format = "0.00"
    for col in range(1, 3):
        cell = ws_an.cell(row=4, column=col)
        cell.border = border
        cell.alignment = right if col == 2 else Alignment(horizontal="left")
        cell = ws_an.cell(row=5, column=col)
        cell.border = border
        cell.alignment = right if col == 2 else Alignment(horizontal="left")
    ws_an.freeze_panes = ws_an["A4"]

    # Return response
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=laporan_keuangan.xlsx'
    return resp

@login_required
def neraca_saldo_pdf(request):
    """Export Neraca Saldo to PDF - Only user's data"""
    # Build neraca data using shared helper so saldo-normal logic matches other reports
    data = compute_account_totals_for_user(request.user)
    neraca_data = []
    for entry in data['saldo']:
        akun = entry['akun']
        debit = entry.get('debit', 0) or 0
        kredit = entry.get('kredit', 0) or 0
        if debit >= kredit:
            disp_debet = debit - kredit
            disp_kredit = 0
        else:
            disp_debet = 0
            disp_kredit = kredit - debit
        neraca_data.append({
            'akun__kode': akun.kode,
            'akun__nama': akun.nama,
            'total_debet': disp_debet,
            'total_kredit': disp_kredit,
        })

    buffer = export_neraca_saldo_pdf(neraca_data)
    response = create_pdf_response('neraca_saldo.pdf')
    response.write(buffer.getvalue())

    return response


@login_required
def laporan_keuangan_pdf(request):
    """Export Laporan Keuangan to PDF - Only user's data"""
    data = compute_account_totals_for_user(request.user)

    pendapatan = []
    beban = []
    aktiva = []
    kewajiban = []
    modal = []

    # Build lists per section
    for entry in data['saldo']:
        akun = entry['akun']
        saldo_akhir = entry['saldo']
        if akun.tipe in ['Aset', 'Aktiva', 'Aktiva Lancar', 'Aktiva Tetap', 'Aktiva Lainnya']:
            aktiva.append({'akun__nama': akun.nama, 'saldo': abs(saldo_akhir) if saldo_akhir > 0 else 0})
        elif akun.tipe in ['Kewajiban', 'Kewajiban Lancar', 'Kewajiban Jangka Panjang']:
            kewajiban.append({'akun__nama': akun.nama, 'saldo': abs(saldo_akhir) if saldo_akhir < 0 else 0})
        elif akun.tipe == 'Modal':
            modal.append({'akun__nama': akun.nama, 'saldo': abs(saldo_akhir) if saldo_akhir < 0 else 0})
        elif akun.tipe == 'Pendapatan':
            total = abs(saldo_akhir) if saldo_akhir < 0 else 0
            pendapatan.append({'akun__nama': akun.nama, 'total': total})
        elif akun.tipe == 'Beban':
            total = abs(saldo_akhir) if saldo_akhir > 0 else 0
            beban.append({'akun__nama': akun.nama, 'total': total})

    total_aktiva = data.get('total_aktiva', 0)
    total_kewajiban = data.get('total_kewajiban', 0)
    total_modal = data.get('total_modal', 0)
    total_pendapatan = data.get('total_pendapatan', 0)
    total_beban = data.get('total_beban', 0)

    # Net income (laba/rugi bersih)
    net_income = total_pendapatan - total_beban
    total_modal_display = total_modal + net_income
    equation_right = total_kewajiban + total_modal_display

    # Simple cash flow reconstruction
    kas_masuk_operasional = total_pendapatan
    kas_keluar_operasional = total_beban
    kas_bersih_operasional = kas_masuk_operasional - kas_keluar_operasional
    kas_masuk_investasi = 0
    kas_keluar_investasi = 0
    kas_bersih_investasi = kas_masuk_investasi - kas_keluar_investasi
    kas_masuk_pendanaan = total_modal
    kas_keluar_pendanaan = 0
    kas_bersih_pendanaan = kas_masuk_pendanaan - kas_keluar_pendanaan
    total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi + kas_bersih_pendanaan

    # Ratios
    profit_margin = (net_income / total_pendapatan * 100) if total_pendapatan else 0
    current_ratio = (total_aktiva / total_kewajiban) if total_kewajiban else None

    laporan_data = {
        'pendapatan': pendapatan,
        'beban': beban,
        'aktiva': aktiva,
        'kewajiban': kewajiban,
        'modal': modal,
        'total_aktiva': total_aktiva,
        'total_kewajiban': total_kewajiban,
        'total_modal_display': total_modal_display,
        'net_income': net_income,
        'equation_right': equation_right,
        'kas_masuk_operasional': kas_masuk_operasional,
        'kas_keluar_operasional': kas_keluar_operasional,
        'kas_bersih_operasional': kas_bersih_operasional,
        'kas_masuk_investasi': kas_masuk_investasi,
        'kas_keluar_investasi': kas_keluar_investasi,
        'kas_bersih_investasi': kas_bersih_investasi,
        'kas_masuk_pendanaan': kas_masuk_pendanaan,
        'kas_keluar_pendanaan': kas_keluar_pendanaan,
        'kas_bersih_pendanaan': kas_bersih_pendanaan,
        'total_kas_bersih': total_kas_bersih,
        'profit_margin': profit_margin,
        'current_ratio': current_ratio,
    }

    buffer = export_laporan_keuangan_pdf(laporan_data)
    response = create_pdf_response('laporan_keuangan.pdf')
    response.write(buffer.getvalue())

    return response
